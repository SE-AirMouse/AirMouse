import numpy
import pyautogui
import mouse
import tools
import copy
import openpyxl



wb = openpyxl.Workbook()
wb.create_sheet("test1")
s1 = wb['test1']            # 開啟工作表

s1.cell(1,1).value = "name"
s1.cell(1,2).value = "Kevin"

wb.save("dataGen/testing.xlsx")

