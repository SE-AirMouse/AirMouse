import time

runTimeRecorder = time.perf_counter()
print("loading dictionary")

import cv2
import mediapipe as mp
import numpy
import openpyxl

import fps
import tools
import draw3DHand
import gesture

MAX_FPS = 60
CAM_NUM = 0
MAX_HANDS_AMOUNT = 1

# 3d圖xyz比例記得調一樣
# 不然手會爛掉
#
#

time.perf_counter()
print("took " + str(time.perf_counter() - runTimeRecorder) + " sec")
runTimeRecorder = time.perf_counter()
print("loading camera")

cap = cv2.VideoCapture(CAM_NUM)

if cap is None or not cap.isOpened():
    raise Exception(
        'Unable to access camera, please check README.md for more info')

print("took " + str(time.perf_counter() - runTimeRecorder) + " sec")
runTimeRecorder = time.perf_counter()
print("loading model")

mpHands = mp.solutions.hands
handModel = mpHands.Hands(model_complexity=1,
                          max_num_hands=MAX_HANDS_AMOUNT,
                          min_detection_confidence=0.95,
                          min_tracking_confidence=0.05)
mpDraw = mp.solutions.drawing_utils

print("took " + str(time.perf_counter() - runTimeRecorder) + " sec")
runTimeRecorder = time.perf_counter()

handLandmarkConnection = numpy.array([
    [0, 1],
    [1, 2],
    [2, 3],
    [3, 4],
    [0, 5],
    [5, 6],
    [6, 7],
    [7, 8],
    [9, 10],
    [10, 11],
    [11, 12],
    [13, 14],
    [14, 15],
    [15, 16],
    [0, 17],
    [17, 18],
    [18, 19],
    [19, 20],
    [5, 9],
    [9, 13],
    [13, 17],
])

hand3D = draw3DHand.handDrawer3D()

FPS = fps.fps()
standardRelativeLandmarkPos = numpy.zeros((21, 3))
standardFingerdegrees = numpy.zeros((5))
startRecord = False
recordRow = 8

wb = openpyxl.Workbook()
wb.create_sheet("row")
s1 = wb['row']  # 開啟工作表

while (True):

    curFps = FPS.get()

    ret, img = cap.read()

    if (ret):
        imgHigh = img.shape[0]
        imgWidth = img.shape[1]
        imgSize = (imgHigh + imgWidth) / 2

        imgRGB = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        handResult = handModel.process(imgRGB)
        if (handResult.multi_hand_landmarks):
            handLandmark = numpy.zeros((21, 3))

            for i in range(21):
                handLandmark[i][0] = (handResult.multi_hand_landmarks[0].
                                      landmark[i].x) * imgWidth / imgSize
                handLandmark[i][1] = (handResult.multi_hand_landmarks[0].
                                      landmark[i].y) * imgHigh / imgSize
                handLandmark[i][2] = (handResult.multi_hand_landmarks[0].
                                      landmark[i].z) * imgWidth / imgSize

            vector_a = handLandmark[13] - handLandmark[0]
            vector_b = handLandmark[13] - handLandmark[5]
            handFaceVector = tools.externalProduct(vector_a, vector_b)
            handFaceVector = handFaceVector

            handFaceYaw = tools.getDegree(
                numpy.array([handFaceVector[0], handFaceVector[2]]),
                numpy.array([0, -1]))
            if (handFaceVector[0] < 0):
                handFaceYaw = -handFaceYaw

            handFacePitch = tools.getDegree(
                numpy.array(
                    [handFaceVector[0], handFaceVector[1], handFaceVector[2]]),
                numpy.array([handFaceVector[0], 0, handFaceVector[2]]))
            if (handFaceVector[1] < 0):
                handFacePitch = -handFacePitch

            handSize = tools.getVectorLength(handLandmark[5] -
                                             handLandmark[17])

            relativeLandmarkPos = numpy.zeros((21, 3))
            for i in range(21):
                relativeLandmarkPos[i] = (handLandmark[i] -
                                          handLandmark[13]) / handSize

                relativeLandmarkPos[i][0], relativeLandmarkPos[i][
                    2] = tools.rotateVector(
                        numpy.array([
                            relativeLandmarkPos[i][0],
                            relativeLandmarkPos[i][2]
                        ]), -handFaceYaw)

                relativeLandmarkPos[i][1], relativeLandmarkPos[i][
                    2] = tools.rotateVector(
                        numpy.array([
                            relativeLandmarkPos[i][1],
                            relativeLandmarkPos[i][2]
                        ]), -handFacePitch)

            handFaceRowVector = relativeLandmarkPos[13] - relativeLandmarkPos[0]
            handFaceRow = tools.getDegree(handFaceRowVector,
                                          numpy.array([0, -1, 0]))
            if (handFaceRowVector[0] < 0):
                handFaceRow = -handFaceRow

            # print(relativeLandmarkPos[13], relativeLandmarkPos[0])

            for i in range(21):
                relativeLandmarkPos[i][0], relativeLandmarkPos[i][
                    1] = tools.rotateVector(
                        numpy.array([
                            relativeLandmarkPos[i][0],
                            relativeLandmarkPos[i][1]
                        ]), -handFaceRow)

            fingerdegrees = gesture.analize(relativeLandmarkPos,
                                            returnDegree=True)

            cv2KeyEvent = cv2.waitKey(1)
            if (cv2KeyEvent == ord('s') and not startRecord):

                print("start record!!!")
                standardRelativeLandmarkPos = relativeLandmarkPos.copy()
                standardFingerdegrees = fingerdegrees.copy()

                s1.cell(1, 1).value = "standerd"

                s1.cell(1, 3).value = "yaw"
                s1.cell(2, 3).value = handFaceYaw
                s1.cell(1, 4).value = "pitch"
                s1.cell(2, 4).value = handFacePitch
                s1.cell(1, 5).value = "row"
                s1.cell(2, 5).value = handFaceRow

                s1.cell(1, 7).value = "thumb"
                s1.cell(2, 7).value = fingerdegrees[0]
                s1.cell(1, 8).value = "fore"
                s1.cell(2, 8).value = fingerdegrees[1]
                s1.cell(1, 9).value = "middle"
                s1.cell(2, 9).value = fingerdegrees[2]
                s1.cell(1, 10).value = "ring"
                s1.cell(2, 10).value = fingerdegrees[3]
                s1.cell(1, 11).value = "little"
                s1.cell(2, 11).value = fingerdegrees[4]

                s1.cell(1, 13).value = "landmark"
                s1.cell(2, 13).value = "x"
                s1.cell(3, 13).value = "y"
                s1.cell(4, 13).value = "z"
                for i in range(21):
                    s1.cell(1, 14 + i).value = str(i)
                    s1.cell(2, 14 + i).value = relativeLandmarkPos[i][0]
                    s1.cell(3, 14 + i).value = relativeLandmarkPos[i][1]
                    s1.cell(4, 14 + i).value = relativeLandmarkPos[i][2]

                s1.cell(7, 1).value = "others"

                s1.cell(7, 3).value = "yaw"
                s1.cell(7, 4).value = "pitch"
                s1.cell(7, 5).value = "row"

                s1.cell(7, 7).value = "thumb"
                s1.cell(7, 8).value = "fore"
                s1.cell(7, 9).value = "middle"
                s1.cell(7, 10).value = "ring"
                s1.cell(7, 11).value = "little"

                s1.cell(7, 13).value = "landmark"
                s1.cell(8, 13).value = "dist"
                for i in range(21):
                    s1.cell(7, 14 + i).value = str(i)

                startRecord = True

            if (cv2KeyEvent == ord('r') and startRecord):
                print("recording!!!")
                s1.cell(recordRow, 3).value = handFaceYaw
                s1.cell(recordRow, 4).value = handFacePitch
                s1.cell(recordRow, 5).value = handFaceRow

                s1.cell(recordRow,
                        7).value = fingerdegrees[0] - standardFingerdegrees[0]
                s1.cell(recordRow,
                        8).value = fingerdegrees[1] - standardFingerdegrees[1]
                s1.cell(recordRow,
                        9).value = fingerdegrees[2] - standardFingerdegrees[2]
                s1.cell(recordRow,
                        10).value = fingerdegrees[3] - standardFingerdegrees[3]
                s1.cell(recordRow,
                        11).value = fingerdegrees[4] - standardFingerdegrees[4]

                for i in range(21):
                    s1.cell(recordRow, 14 + i).value = tools.getVectorLength(
                        relativeLandmarkPos[i] -
                        standardRelativeLandmarkPos[i])

                recordRow += 1

            if (cv2KeyEvent == ord('a') and startRecord):
                recordRow += 1
            # print(handFaceVector)
            print(handFaceYaw, handFacePitch, handFaceRow)

            # print(relativeLandmarkPos[4])
            hand3D.draw(relativeLandmarkPos)

            for handLms in handResult.multi_hand_landmarks:
                mpDraw.draw_landmarks(img, handLms, mpHands.HAND_CONNECTIONS)

        cv2.putText(img, "fps: " + str(int(curFps)), (0, 50),
                    cv2.FONT_HERSHEY_SIMPLEX, 2, (0, 0, 0), 2)
        cv2.imshow("cam", img)
    else:
        pass

    cv2KeyEvent = cv2.waitKey(1)
    if (cv2KeyEvent == ord('q')):
        break

    if (cv2KeyEvent == 27):
        break

wb.save('dataGen/row-掌.xlsx')
