import numpy
import openpyxl
import tools
import gesture


class modelAccuracyRecorder:
    def __init__(self, name="new data") -> None:
        self.dataName = name
        self.wb = openpyxl.Workbook()
        self.wb.create_sheet(self.dataName)
        self.s1 = self.wb[self.dataName]
        self.standardRelativeLandmark = numpy.zeros((21, 3))
        self.standardSet = False
        self.recordRow = 8

    def setStandard(self, landmark) -> None:
        # input the standard xyz 1:1:1 scale hand landmark position

        relativeHand = self.__relativeHand(landmark)
        relativeLandmark = relativeHand["landmark"]
        self.standardRelativeLandmark = relativeLandmark
        fingerdegrees = gesture.analize(relativeLandmark, returnDegree=True)
        self.standardFingerdegrees = fingerdegrees

        self.s1.cell(1, 1).value = "standerd"

        self.s1.cell(1, 3).value = "yaw"
        self.s1.cell(2, 3).value = relativeHand["yaw"]
        self.s1.cell(1, 4).value = "pitch"
        self.s1.cell(2, 4).value = relativeHand["pitch"]
        self.s1.cell(1, 5).value = "roll"
        self.s1.cell(2, 5).value = relativeHand["roll"]

        self.s1.cell(1, 7).value = "fingers"
        self.s1.cell(2, 7).value = "original degrees"
        self.s1.cell(1, 8).value = "thumb"
        self.s1.cell(2, 8).value = fingerdegrees[0]
        self.s1.cell(1, 9).value = "fore"
        self.s1.cell(2, 9).value = fingerdegrees[1]
        self.s1.cell(1, 10).value = "middle"
        self.s1.cell(2, 10).value = fingerdegrees[2]
        self.s1.cell(1, 11).value = "ring"
        self.s1.cell(2, 11).value = fingerdegrees[3]
        self.s1.cell(1, 12).value = "little"
        self.s1.cell(2, 12).value = fingerdegrees[4]

        self.s1.cell(1, 14).value = "landmark"
        self.s1.cell(2, 14).value = "x"
        self.s1.cell(3, 14).value = "y"
        self.s1.cell(4, 14).value = "z"
        for i in range(21):
            self.s1.cell(1, 15 + i).value = str(i)
            self.s1.cell(2, 15 + i).value = relativeLandmark[i][0]
            self.s1.cell(3, 15 + i).value = relativeLandmark[i][1]
            self.s1.cell(4, 15 + i).value = relativeLandmark[i][2]

        self.s1.cell(7, 1).value = "others"

        self.s1.cell(7, 3).value = "yaw"
        self.s1.cell(7, 4).value = "pitch"
        self.s1.cell(7, 5).value = "roll"

        self.s1.cell(7, 7).value = "fingers"
        self.s1.cell(8, 7).value = "diff"
        self.s1.cell(7, 8).value = "thumb"
        self.s1.cell(7, 9).value = "fore"
        self.s1.cell(7, 10).value = "middle"
        self.s1.cell(7, 11).value = "ring"
        self.s1.cell(7, 12).value = "little"

        self.s1.cell(7, 14).value = "landmark"
        self.s1.cell(8, 15).value = "dist"
        for i in range(21):
            self.s1.cell(7, 15 + i).value = str(i)

        self.standardSet = True

    def record(self, landmark) -> None:
        # must run self.setStandard before running this
        # input the other xyz 1:1:1 scale hand landmark position

        if not self.standardSet:
            raise RuntimeError(
                "didn't set data standard before recording relative data"
            )

        relativeHand = self.__relativeHand(landmark)
        relativeLandmark = relativeHand["landmark"]
        fingerdegrees = gesture.analize(relativeLandmark, returnDegree=True)

        self.s1.cell(self.recordRow, 3).value = relativeHand["yaw"]
        self.s1.cell(self.recordRow, 4).value = relativeHand["pitch"]
        self.s1.cell(self.recordRow, 5).value = relativeHand["roll"]

        self.s1.cell(self.recordRow, 8).value = (
            fingerdegrees[0] - self.standardFingerdegrees[0]
        )
        self.s1.cell(self.recordRow, 9).value = (
            fingerdegrees[1] - self.standardFingerdegrees[1]
        )
        self.s1.cell(self.recordRow, 10).value = (
            fingerdegrees[2] - self.standardFingerdegrees[2]
        )
        self.s1.cell(self.recordRow, 11).value = (
            fingerdegrees[3] - self.standardFingerdegrees[3]
        )
        self.s1.cell(self.recordRow, 12).value = (
            fingerdegrees[4] - self.standardFingerdegrees[4]
        )

        for i in range(21):
            self.s1.cell(self.recordRow, 15 + i).value = tools.getVectorLength(
                relativeLandmark[i] - self.standardRelativeLandmark[i]
            )

        self.recordRow += 1

    def push(self, landmark) -> None:
        if not self.standardSet:
            self.setStandard(landmark)
        else:
            self.record(landmark)

    def skipRow(self, row=1) -> None:
        self.record += row

    def __relativeHand(self, landmark) -> dict:
        vector_a = landmark[13] - landmark[0]
        vector_b = landmark[13] - landmark[5]
        handFaceVector = tools.externalProduct(vector_a, vector_b)
        handFaceVector = handFaceVector

        handFaceYaw = tools.getDegree(
            numpy.array([handFaceVector[0], handFaceVector[2]]),
            numpy.array([0, -1]),
        )
        if handFaceVector[0] < 0:
            handFaceYaw = -handFaceYaw

        handFacePitch = tools.getDegree(
            numpy.array([handFaceVector[0], handFaceVector[1], handFaceVector[2]]),
            numpy.array([handFaceVector[0], 0, handFaceVector[2]]),
        )
        if handFaceVector[1] < 0:
            handFacePitch = -handFacePitch

        handSize = tools.getVectorLength(landmark[5] - landmark[17])

        relativeLandmark = numpy.zeros((21, 3))
        for i in range(21):
            relativeLandmark[i] = (landmark[i] - landmark[13]) / handSize

            relativeLandmark[i][0], relativeLandmark[i][2] = tools.rotateVector(
                numpy.array([relativeLandmark[i][0], relativeLandmark[i][2]]),
                -handFaceYaw,
            )

            relativeLandmark[i][1], relativeLandmark[i][2] = tools.rotateVector(
                numpy.array([relativeLandmark[i][1], relativeLandmark[i][2]]),
                -handFacePitch,
            )

        handFaceRollVector = relativeLandmark[13] - relativeLandmark[0]
        handFaceRoll = tools.getDegree(handFaceRollVector, numpy.array([0, -1, 0]))
        if handFaceRollVector[0] < 0:
            handFaceRoll = -handFaceRoll

        for i in range(21):
            relativeLandmark[i][0], relativeLandmark[i][1] = tools.rotateVector(
                numpy.array([relativeLandmark[i][0], relativeLandmark[i][1]]),
                -handFaceRoll,
            )

        result = self.actionStatus = {
            "landmark": relativeLandmark,
            "yaw": handFaceYaw,
            "pitch": handFacePitch,
            "roll": handFaceRoll,
        }

        return result

    def finish(self) -> None:
        self.wb.save("dataGen/" + self.dataName + ".xlsx")
